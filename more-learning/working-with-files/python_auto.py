#working with spreadsheet files
import openpyxl


file = openpyxl.load_workbook("inventory.xlsx")
sheet_one = file["Sheet1"]

product_per_supplier = {}
amount_per_supplier = {}
products_under_ten = {}

for sheet_row in range(2, sheet_one.max_row + 1):
    supplier = sheet_one.cell(sheet_row, 4).value
    inventory = sheet_one.cell(sheet_row, 2).value
    price = sheet_one.cell(sheet_row, 3).value
    product_no = sheet_one.cell(sheet_row, 1).value
    inventory_price = sheet_one.cell(sheet_row, 5)

    #number of products per supplier
    if supplier in product_per_supplier:
        product_per_supplier[supplier] = product_per_supplier[supplier] + 1
    else:
        product_per_supplier[supplier] = 1  
    
    #total amount of inventory per supplier
    if supplier in amount_per_supplier:
        current_total = amount_per_supplier[supplier]
        amount_per_supplier[supplier] = current_total + inventory * price
    else:
        amount_per_supplier[supplier] = inventory * price

    #supplier with inventory under 10
    if inventory < 10:
        products_under_ten[int(product_no)] = int(inventory)

    #adding total amount
    inventory_price.value = inventory * price

print(product_per_supplier)
print(amount_per_supplier)
print(products_under_ten)

file.save("inventory-w-total-amount.xlsx")